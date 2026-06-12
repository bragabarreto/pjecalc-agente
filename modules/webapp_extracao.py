"""Extração IA in-app — Fase 1 (12/06/2026).

Replica o fluxo do Projeto Claude externo DENTRO do aplicativo:
o usuário cola o texto da sentença (ou sobe PDF/DOCX/TXT/MD) e anexa
documentos do processo (PDF, imagens, MD/TXT, planilhas XLSX); a extração
roda via API Anthropic com o MESMO prompt do projeto externo
(`SYSTEM_PROMPT_V2_EXTERNAL` — fonte única em modules/extraction_v2.py)
e o MESMO fluxo de 2 etapas:

    Etapa 1: resumo prévio em markdown → usuário revisa/corrige na tela
    Etapa 2: após "Confirmar", JSON v2 → normalizer → prévia v2

A partir da Etapa 2 o fluxo desemboca 100% no pipeline v2 existente
(`_save_previa` → /previa/v2/{id} → automação) — nada do v2 é alterado.

⚠ PRESERVAÇÃO (regra do CLAUDE.md): este módulo é ADITIVO. Os caminhos
existentes — colar/subir JSON do projeto externo via /processar/v2 ou
auto-detecção .json no /processar — permanecem intocados e continuam
sendo opção de entrada. Rotas novas usam o sufixo /ia para não colidir
com /previa_v3 (UI antiga da prévia v1).

Regra IA-only: qualquer falha da API Anthropic → fase "erro" com mensagem
clara. NUNCA gerar prévia por fallback regex.
"""

from __future__ import annotations

import base64
import json
import logging
import os
import re
import threading
import time
import uuid
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates

logger = logging.getLogger(__name__)

router_extracao = APIRouter()
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))

# ─── Storage de sessões de extração (mesma cascata do webapp_v2) ──────────
_CANDIDATOS = [
    os.environ.get("EXTRACAO_IA_DIR"),
    "/app/data/calculations/extracao_ia",
    "data/calculations/extracao_ia",
    "/tmp/pjecalc_extracao_ia",
]
_STORE_DIR: Path | None = None
for _c in _CANDIDATOS:
    if not _c:
        continue
    try:
        _p = Path(_c)
        _p.mkdir(parents=True, exist_ok=True)
        _t = _p / ".write_test"
        _t.write_text("ok")
        _t.unlink()
        _STORE_DIR = _p
        break
    except Exception:
        continue
if _STORE_DIR is None:
    _STORE_DIR = Path("/tmp/pjecalc_extracao_ia")
    _STORE_DIR.mkdir(parents=True, exist_ok=True)
logger.info(f"webapp_extracao: _STORE_DIR = {_STORE_DIR}")

_MAX_FILE_BYTES = 15 * 1024 * 1024
_IMG_MAX_PX = 1536
_MAX_EXTRAS = 10
_MAX_TOKENS_ETAPA1 = 16000
_MAX_TOKENS_ETAPA2 = 32000


def _sessao_dir(sessao_id: str) -> Path:
    d = _STORE_DIR / sessao_id
    (d / "files").mkdir(parents=True, exist_ok=True)
    return d


def _load_estado(sessao_id: str) -> dict | None:
    f = _STORE_DIR / sessao_id / "estado.json"
    if not f.exists():
        return None
    return json.loads(f.read_text(encoding="utf-8"))


def _save_estado(sessao_id: str, estado: dict) -> None:
    f = _STORE_DIR / sessao_id / "estado.json"
    f.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


# ─── Conversão de arquivos para blocos de conteúdo da API ─────────────────


def _xlsx_para_markdown(path: Path, max_linhas: int = 200) -> str:
    """Converte planilha em tabelas markdown (uma por aba)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    partes: list[str] = []
    for ws in wb.worksheets:
        linhas = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_linhas:
                linhas.append(f"... ({ws.max_row - max_linhas} linhas omitidas)")
                break
            cels = ["" if c is None else str(c) for c in row]
            linhas.append("| " + " | ".join(cels) + " |")
            if i == 0:
                linhas.append("|" + "---|" * len(cels))
        if linhas:
            partes.append(f"### Aba: {ws.title}\n" + "\n".join(linhas))
    wb.close()
    return "\n\n".join(partes) or "(planilha vazia)"


def _arquivo_para_bloco(meta: dict) -> list[dict]:
    """Converte um arquivo salvo em bloco(s) de conteúdo da API Anthropic.

    Tipos: pdf → document block; imagem → image block;
    docx/txt/md → texto extraído; xlsx → tabela markdown.
    """
    path = Path(meta["caminho"])
    suf = path.suffix.lower()
    contexto = meta.get("contexto") or ""
    prefixo = f"=== DOCUMENTO: {meta.get('nome', path.name)}"
    if contexto:
        prefixo += f" ({contexto})"
    prefixo += " ==="

    if meta["tipo"] == "imagem":
        data = base64.standard_b64encode(path.read_bytes()).decode()
        blocos: list[dict] = [{"type": "text", "text": prefixo}]
        blocos.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": meta.get("mime_type", "image/jpeg"),
                "data": data,
            },
        })
        return blocos

    if suf == ".pdf":
        data = base64.standard_b64encode(path.read_bytes()).decode()
        return [
            {"type": "text", "text": prefixo},
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": data,
                },
            },
        ]

    if suf in (".docx", ".doc"):
        from modules.ingestion import _ler_docx

        res = _ler_docx(path, [])
        return [{"type": "text", "text": f"{prefixo}\n{res.get('texto', '')[:50000]}"}]

    if suf in (".xlsx", ".xlsm"):
        try:
            md = _xlsx_para_markdown(path)
        except Exception as e:
            md = f"(falha ao ler planilha: {e})"
        return [{"type": "text", "text": f"{prefixo}\n{md[:50000]}"}]

    # txt / md / csv / fallback texto
    try:
        txt = path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        txt = f"(falha ao ler arquivo: {e})"
    return [{"type": "text", "text": f"{prefixo}\n{txt[:50000]}"}]


def _montar_conteudo_etapa1(estado: dict) -> list[dict]:
    """Remonta os blocos da 1ª mensagem do usuário a partir do estado."""
    blocos: list[dict] = []
    texto = (estado.get("texto_colado") or "").strip()
    if texto:
        blocos.append({
            "type": "text",
            "text": f"=== SENTENÇA (texto colado) ===\n{texto}",
        })
    for meta in estado.get("arquivos", []):
        try:
            blocos.extend(_arquivo_para_bloco(meta))
        except Exception as e:
            logger.warning(f"bloco de {meta.get('nome')}: {e}")
            blocos.append({
                "type": "text",
                "text": f"(documento {meta.get('nome')} não pôde ser lido: {e})",
            })
    blocos.append({
        "type": "text",
        "text": (
            "Execute a ETAPA 1 do fluxo operacional: apresente o resumo "
            "prévio em markdown para minha validação. NÃO gere o JSON ainda."
        ),
    })
    return blocos


def _montar_messages(estado: dict, nova_msg: str | None = None) -> list[dict]:
    """Histórico completo da conversa para a API (stateless).

    Prompt caching: o último bloco da 1ª mensagem recebe cache_control —
    isso cacheia o PREFIXO inteiro (system prompt ~25k tokens + sentença +
    documentos). Etapa 1 grava o cache; correções e Etapa 2 (minutos
    depois, mesmo prefixo) leem com 90% de desconto na entrada.
    """
    blocos = _montar_conteudo_etapa1(estado)
    if blocos:
        blocos[-1]["cache_control"] = {"type": "ephemeral"}
    msgs: list[dict] = [{"role": "user", "content": blocos}]
    for turno in estado.get("conversa", []):
        msgs.append({"role": turno["role"], "content": turno["texto"]})
    if nova_msg:
        msgs.append({"role": "user", "content": nova_msg})
    return msgs


def _chamar_claude(messages: list[dict], max_tokens: int) -> str:
    import anthropic

    from config import ANTHROPIC_API_KEY, CLAUDE_MODEL
    from modules.extraction_v2 import SYSTEM_PROMPT_V2_EXTERNAL

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=600.0)
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=max_tokens,
        # cache_control no system: o prompt (~25k tokens) é idêntico em
        # TODA chamada — a 1ª grava o cache, as demais pagam 10% na leitura
        system=[{
            "type": "text",
            "text": SYSTEM_PROMPT_V2_EXTERNAL,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=messages,
        temperature=0.0,
    )
    u = getattr(resp, "usage", None)
    if u is not None:
        logger.info(
            "extracao-ia tokens: in=%s out=%s cache_write=%s cache_read=%s",
            getattr(u, "input_tokens", "?"),
            getattr(u, "output_tokens", "?"),
            getattr(u, "cache_creation_input_tokens", 0),
            getattr(u, "cache_read_input_tokens", 0),
        )
    return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()


def _extrair_json(texto: str) -> dict:
    """Extrai o objeto JSON da resposta da Etapa 2 (tolerante a fences)."""
    t = texto.strip()
    if t.startswith("```"):
        t = re.sub(r"^```[a-zA-Z]*\n", "", t)
        t = re.sub(r"\n```\s*$", "", t)
    try:
        return json.loads(t)
    except json.JSONDecodeError:
        # fallback: maior bloco {...} da resposta
        ini, fim = t.find("{"), t.rfind("}")
        if ini >= 0 and fim > ini:
            return json.loads(t[ini : fim + 1])
        raise


# ─── Workers em background (thread própria — chamadas longas à API) ───────


def _worker_etapa1(sessao_id: str) -> None:
    estado = _load_estado(sessao_id) or {}
    try:
        resumo = _chamar_claude(_montar_messages(estado), _MAX_TOKENS_ETAPA1)
        estado["conversa"] = estado.get("conversa", [])
        estado["conversa"].append({"role": "assistant", "texto": resumo})
        estado["resumo_md"] = resumo
        estado["fase"] = "resumo_pronto"
    except Exception as e:
        logger.exception(f"[{sessao_id}] etapa 1 falhou")
        estado["fase"] = "erro"
        estado["erro"] = f"Falha na extração via IA (Etapa 1): {e}"
    _save_estado(sessao_id, estado)


def _worker_correcao(sessao_id: str, correcoes: str) -> None:
    estado = _load_estado(sessao_id) or {}
    try:
        estado.setdefault("conversa", []).append({"role": "user", "texto": correcoes})
        _save_estado(sessao_id, estado)
        resumo = _chamar_claude(_montar_messages(estado), _MAX_TOKENS_ETAPA1)
        estado["conversa"].append({"role": "assistant", "texto": resumo})
        estado["resumo_md"] = resumo
        estado["fase"] = "resumo_pronto"
    except Exception as e:
        logger.exception(f"[{sessao_id}] correção falhou")
        estado["fase"] = "erro"
        estado["erro"] = f"Falha ao aplicar correções via IA: {e}"
    _save_estado(sessao_id, estado)


def _worker_etapa2(sessao_id: str) -> None:
    estado = _load_estado(sessao_id) or {}
    try:
        estado.setdefault("conversa", []).append({"role": "user", "texto": "confirmar"})
        _save_estado(sessao_id, estado)
        bruto = _chamar_claude(_montar_messages(estado), _MAX_TOKENS_ETAPA2)
        estado["conversa"].append({"role": "assistant", "texto": bruto})
        payload = _extrair_json(bruto)

        # mesmo pipeline do /processar/v2: normalizer → Pydantic → store v2
        from modules.json_normalizer import normalize_v2_json
        from modules.webapp_v2 import PreviaCalculoV2, _save_previa

        payload = normalize_v2_json(payload)
        previa = PreviaCalculoV2.model_validate(payload)
        _save_previa(sessao_id, previa.model_dump())

        estado["fase"] = "previa_pronta"
        estado["url_previa"] = f"/previa/v2/{sessao_id}"
    except Exception as e:
        logger.exception(f"[{sessao_id}] etapa 2 falhou")
        # volta para resumo_pronto: o usuário pode corrigir e tentar de novo
        estado["fase"] = "erro_etapa2"
        estado["erro"] = f"Falha na geração do JSON (Etapa 2): {e}"
    _save_estado(sessao_id, estado)


def _disparar(worker, *args) -> None:
    threading.Thread(target=worker, args=args, daemon=True).start()


# ─── Rotas ────────────────────────────────────────────────────────────────


@router_extracao.get("/novo/ia", response_class=HTMLResponse)
async def pagina_novo_ia(request: Request):
    """Formulário: colar sentença ou subir arquivo + documentos extras."""
    return templates.TemplateResponse(request, "novo_calculo_ia.html", {})


@router_extracao.post("/processar/ia")
async def processar_ia(request: Request):
    """Recebe sentença (texto colado e/ou arquivo) + extras; dispara Etapa 1."""
    form = await request.form()
    sessao_id = str(uuid.uuid4())
    sdir = _sessao_dir(sessao_id)

    texto_colado = str(form.get("texto_sentenca", "")).strip()
    arquivos: list[dict] = []

    async def _salvar_upload(up, idx: int, contexto: str, eh_imagem: bool) -> None:
        dados = await up.read()
        if not dados or len(dados) > _MAX_FILE_BYTES:
            return
        nome = Path(up.filename).name
        suf = Path(nome).suffix.lower()
        meta: dict = {"nome": nome, "contexto": contexto}
        if eh_imagem or suf in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
            import io

            from PIL import Image as _PIL

            try:
                img = _PIL.open(io.BytesIO(dados))
                img.thumbnail((_IMG_MAX_PX, _IMG_MAX_PX), _PIL.Resampling.LANCZOS)
                if img.mode == "RGBA":
                    img = img.convert("RGB")
                buf = io.BytesIO()
                fmt = "JPEG" if img.mode in ("RGB", "L") else "PNG"
                img.save(buf, format=fmt, quality=85)
                dados = buf.getvalue()
                meta["mime_type"] = "image/jpeg" if fmt == "JPEG" else "image/png"
            except Exception:
                meta["mime_type"] = getattr(up, "content_type", "image/jpeg")
            meta["tipo"] = "imagem"
            path = sdir / "files" / f"{idx}_{Path(nome).stem}.jpg"
        else:
            meta["tipo"] = "arquivo"
            path = sdir / "files" / f"{idx}_{nome}"
        path.write_bytes(dados)
        meta["caminho"] = str(path)
        arquivos.append(meta)

    # sentença em arquivo (opcional — texto colado também vale)
    sent = form.get("arquivo_sentenca")
    if sent is not None and hasattr(sent, "filename") and sent.filename:
        await _salvar_upload(sent, 0, "sentença/decisão principal", eh_imagem=False)

    # documentos extras (mesmo padrão do form v1: doc_arquivo_N / doc_imagem_N
    # / doc_texto_N / doc_contexto_N)
    for i in range(_MAX_EXTRAS):
        ctx = str(form.get(f"doc_contexto_{i}", "")).strip()
        arq = form.get(f"doc_arquivo_{i}")
        img = form.get(f"doc_imagem_{i}")
        txt = form.get(f"doc_texto_{i}")
        if arq is not None and hasattr(arq, "filename") and arq.filename:
            await _salvar_upload(arq, i + 1, ctx, eh_imagem=False)
        elif img is not None and hasattr(img, "filename") and img.filename:
            await _salvar_upload(img, i + 1, ctx, eh_imagem=True)
        elif txt and str(txt).strip():
            texto_colado += (
                f"\n\n=== DOCUMENTO COLADO{f' ({ctx})' if ctx else ''} ===\n"
                + str(txt).strip()[:30000]
            )

    if not texto_colado and not arquivos:
        return JSONResponse(
            status_code=400,
            content={"erro": "Cole o texto da sentença ou anexe ao menos um arquivo."},
        )

    estado = {
        "fase": "etapa1_processando",
        "criado_em": time.strftime("%Y-%m-%d %H:%M:%S"),
        "texto_colado": texto_colado,
        "arquivos": arquivos,
        "conversa": [],
    }
    _save_estado(sessao_id, estado)
    _disparar(_worker_etapa1, sessao_id)

    return JSONResponse({
        "sessao_id": sessao_id,
        "url_resumo": f"/resumo/ia/{sessao_id}",
    })


@router_extracao.get("/resumo/ia/{sessao_id}", response_class=HTMLResponse)
async def pagina_resumo_ia(sessao_id: str, request: Request):
    """Página do resumo de validação (Etapa 1) com polling."""
    estado = _load_estado(sessao_id)
    if estado is None:
        return HTMLResponse("Sessão de extração não encontrada.", status_code=404)
    return templates.TemplateResponse(
        request, "resumo_ia.html", {"sessao_id": sessao_id}
    )


@router_extracao.get("/api/ia/{sessao_id}/estado")
async def estado_ia(sessao_id: str):
    estado = _load_estado(sessao_id)
    if estado is None:
        return JSONResponse({"fase": "nao_encontrada"}, status_code=404)
    return JSONResponse({
        "fase": estado.get("fase"),
        "resumo_md": estado.get("resumo_md"),
        "erro": estado.get("erro"),
        "url_previa": estado.get("url_previa"),
        "n_arquivos": len(estado.get("arquivos", [])),
    })


@router_extracao.post("/api/ia/{sessao_id}/corrigir")
async def corrigir_ia(sessao_id: str, payload: dict):
    estado = _load_estado(sessao_id)
    if estado is None:
        return JSONResponse({"erro": "sessão não encontrada"}, status_code=404)
    correcoes = str(payload.get("correcoes", "")).strip()
    if not correcoes:
        return JSONResponse({"erro": "descreva as correções"}, status_code=400)
    if estado.get("fase") not in ("resumo_pronto", "erro_etapa2"):
        return JSONResponse({"erro": f"fase atual: {estado.get('fase')}"}, status_code=409)
    estado["fase"] = "etapa1_processando"
    estado.pop("erro", None)
    _save_estado(sessao_id, estado)
    _disparar(_worker_correcao, sessao_id, correcoes)
    return JSONResponse({"status": "processando"})


@router_extracao.post("/api/ia/{sessao_id}/confirmar")
async def confirmar_ia(sessao_id: str):
    estado = _load_estado(sessao_id)
    if estado is None:
        return JSONResponse({"erro": "sessão não encontrada"}, status_code=404)
    if estado.get("fase") not in ("resumo_pronto", "erro_etapa2"):
        return JSONResponse({"erro": f"fase atual: {estado.get('fase')}"}, status_code=409)
    estado["fase"] = "etapa2_processando"
    estado.pop("erro", None)
    _save_estado(sessao_id, estado)
    _disparar(_worker_etapa2, sessao_id)
    return JSONResponse({"status": "processando"})
